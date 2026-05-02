import csv
from pathlib import Path
from typing import Optional, Set, List, LiteralString, Tuple

from openpyxl import load_workbook

from xls2csv.exception import (
    BatchProcessingError,
    NoSheetFoundError,
    NotAnExcelFileError,
    OutputFileExistsError,
    XLS2CSVError,
)
from xls2csv.utils import (
    DEFAULT_TEMPLATE,
    PathLike,
    format_output_name,
    print_summary,
    is_excel_file,
    iter_excels,
)

# .xls is not supported due to legacy format and openpyxl limitation
SUPPORTED_EXTS: Set[LiteralString] = { ".xlsx", ".xlsb", ".xlsm" }

def convert_single(
    excel_file: PathLike,
    output: Optional[PathLike] = None,
    /,
    *,
    sheet: Optional[str] = None,
    all_sheets: bool = False,
    template: Optional[str] = None,
    overwrite: bool = False
) -> None:
    """
    Convert a single Excel file to CSV format.

    Args:
        excel_file (str or Path-like object): Path to the Excel file to convert.
        output (str or Path-like object, optional): Path to the output CSV file.
        sheet (str, optional): Name of the sheet to convert. Default to active sheet.
        all_sheets (bool, optional): Whether to convert all sheets. Default to active sheet.
        template (str, optional): Template for the output file name. Default to `"%(name)-[%(sheet)].%(ext)"`.
        overwrite (bool, optional): Whether to overwrite existing files. Default to False.

    Raises:
        FileNotFoundError: If the Excel file is not found.
        NoSheetFoundError: If the Excel file does not contain any sheets.
        NotAnExcelFileError: If the given file is not an Excel file.
        OutputFileExistsError: If the output file already exists and overwrite is False.
        XLS2CSVError: If a single output CSV path cannot be used with multiple sheets.

    Note:
    - If `output` is a folder, the CSV file will be saved in that folder.
    - If `output` is a file, the CSV file will be saved in that file.
    - If `output` is not specified, the CSV file will be saved in the same folder as the Excel file.
    - If `all_sheets` is set to True, all sheets will be converted to CSV files.
    - If both `all_sheets` and `sheet` are specified, `all_sheets` will take precedence over sheet.
    """
    excel_file = Path(excel_file)

    if not excel_file.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_file}")

    if not is_excel_file(excel_file, SUPPORTED_EXTS):
        raise NotAnExcelFileError("Given file is not an Excel files", filepath=excel_file)

    wb = load_workbook(excel_file, data_only=True, read_only=True)
    try:
        if all_sheets:  # Take priority over sheet
            sheets = wb.sheetnames
        elif sheet:
            if sheet not in wb.sheetnames:
                raise NoSheetFoundError(sheet, filepath=excel_file)
            sheets = [sheet]
        else:
            if not wb.sheetnames:
                # For edge case; typically, there's no Excel file without a sheet inside
                raise NotAnExcelFileError(f"No sheets found in {excel_file.name!r}")
            sheets = [wb.active.title if wb.active else wb.sheetnames[0]]

        output_path = Path(output) if output is not None else None
        output_is_file = output_path and output_path.suffix.lower() == ".csv"

        if output_is_file and len(sheets) > 1:
            raise XLS2CSVError(
                "A single output CSV path cannot be used with multiple sheets",
                filepath=excel_file,
                err_code="E_IO"
            )

        template = template or DEFAULT_TEMPLATE

        if output_path and not output_is_file:
            output_path.mkdir(parents=True, exist_ok=True)

        for sheet_name in sheets:
            ws = wb[sheet_name]

            # --- filename generation ---
            filename = format_output_name(
                template,
                file=excel_file,
                sheet=sheet_name,
                ext="csv"
            )

            # --- resolve target path ---
            if output_path is None:
                target = excel_file.parent / filename
            elif output_is_file:
                target = output_path
            else:
                target = output_path / filename

            if target.exists() and not overwrite:
                raise OutputFileExistsError(filepath=target)

            with target.open("w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                total_rows: int = 0
                for row in ws.iter_rows(values_only=True):
                    writer.writerow(row)
                    total_rows += 1

            print_summary(excel_file, target, sheet_name, total_rows)
    finally:
        wb.close()

def convert_batch(
    folder: PathLike,
    output: Optional[PathLike] = None,
    /,
    *,
    sheet: Optional[str] = None,
    all_sheets: bool = False,
    template: Optional[str] = None,
    overwrite: bool = False
) -> None:
    """
    Convert all Excel files in a folder to CSV format.

    Args:
        folder (str or Path-like object): Path to the folder containing Excel files.
        output (str or Path-like object, optional): Path to the output folder. Default to the same folder as the Excel files.
        sheet (str, optional): Name of the sheet to convert. Default to active sheet.
        all_sheets (bool, optional): Whether to convert all sheets. Default to active sheet.
        template (str, optional): Template for the output file name. Default to `"%(name)-[%(sheet)].%(ext)"`.
        overwrite (bool, optional): Whether to overwrite existing files. Default to False.

    Raises:
        FileNotFoundError: If the folder is not found.
        NotADirectoryError: If the given path is not a directory.
        ValueError: If no Excel files are found in the folder.
        BatchProcessingError: If any of the Excel files fail to convert.

    Note:
    - If `output` is a folder, the CSV file will be saved in that folder.
    - If `output` is not specified, the CSV file will be saved in the same folder.
    - If `all_sheets` is set to True, all sheets will be converted to CSV files.
    - If both `all_sheets` and `sheet` are specified, `all_sheets` will take precedence over sheet.
    """
    folder = Path(folder)
    premature_err = BatchProcessingError("Premature exit (precondition not met)")

    if not folder.exists():
        premature_err.add_error(FileNotFoundError(f"Folder not found: {folder}"))
        raise premature_err
    if not folder.is_dir():
        premature_err.add_error(NotADirectoryError(f"Expected a directory, got: {folder}"))
        raise premature_err

    output_path = Path(output) if output else folder

    if output_path.exists() and not output_path.is_dir():
        premature_err.add_error(NotADirectoryError("Batch output must be a directory"))
        raise premature_err

    output_path.mkdir(parents=True, exist_ok=True)

    excel_files = list(iter_excels(folder, SUPPORTED_EXTS))

    if not excel_files:
        premature_err.add_error(ValueError(f"No Excel files found in: {folder}"))
        raise premature_err

    errors: List[Tuple[Path, RuntimeError]] = []

    for excel_file in excel_files:
        try:
            print(f"Processing: {excel_file} ({excel_file.stat().st_size / 1024:.2f} KB)")
            convert_single(
                excel_file,
                output_path,
                sheet=sheet,
                all_sheets=all_sheets,
                template=template,
                overwrite=overwrite
            )
        except RuntimeError as e:
            errors.append((excel_file, e))
        finally:
            print("-" * 55)

    if errors:
        err = BatchProcessingError("Batch conversion failed")
        for _, e in errors:
            err.add_error(e)
        raise err

    print(f"\n✅ {len(excel_files)} Excel file(s) converted successfully!")
